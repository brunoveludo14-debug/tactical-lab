"""
SaoMancos Data Sync — Liga a app ao Dashboard Excel
Run: python sync_server.py

O que faz:
1. Abre o Dashboard Excel
2. Recebe dados da app (POST) ou importa CSV
3. Escreve na sheet "Base de Dados"
4. Guarda o ficheiro

Uso:
  - Automático: a app envia POST para http://localhost:5557/sync
  - Manual: python sync_server.py --import ficheiro.csv
"""

from http.server import HTTPServer, BaseHTTPRequestHandler
import json
import csv
import sys
import os
from datetime import datetime
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\bruno\Desktop\São manços\Dashboard_Tactical_Lab.xlsx"

class SyncHandler(BaseHTTPRequestHandler):
    def do_POST(self):
        if self.path == "/sync":
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length)
            data = json.loads(body)

            logs = data.get("logs", [])
            opp_logs = data.get("opp_logs", [])

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.send_response(200)
            self.end_headers()
            resp = {"ok": True, "records": len(logs), "opp": len(opp_logs)}
            self.wfile.write(json.dumps(resp).encode())

            write_to_excel(logs, opp_logs)
            print(f"✅ Sync: {len(logs)} ações + {len(opp_logs)} opponent")
        else:
            self.send_response(404)

    def do_GET(self):
        if self.path == "/status":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"status": "ok", "excel": EXCEL_PATH}).encode())
        else:
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.end_headers()
            self.wfile.write(b"""
            <html><body>
            <h1>Sao Manços Sync Server</h1>
            <p>POST /sync com JSON: {logs: [...], opp_logs: [...]} </p>
            <p>GET /import?file=path.csv para importar CSV</p>
            </body></html>
            """)

    def do_GET_import(self):
        # Handle ?file= query
        pass

def write_to_excel(logs, opp_logs):
    """Escreve os logs na sheet Base de Dados do Excel"""
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel não encontrado: {EXCEL_PATH}")
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Base de Dados"]

    # Find last row with data
    last_row = ws.max_row
    while last_row > 1 and ws.cell(last_row, 1).value is None:
        last_row -= 1

    # Clear old data (keep header row 1)
    for row in range(2, last_row + 1):
        for col in range(1, 7):
            ws.cell(row, col).value = None

    # Write headers if needed
    headers = ["Data", "Minuto", "Jogador", "Numero", "Acao", "Zona"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i).value = h

    # Write new data
    date_str = datetime.now().strftime("%d/%m/%Y")
    for i, log in enumerate(logs, start=2):
        ws.cell(i, 1).value = date_str
        ws.cell(i, 2).value = log.get("min", 0)
        ws.cell(i, 3).value = log.get("player", "")
        ws.cell(i, 4).value = log.get("num", "")
        ws.cell(i, 5).value = log.get("action", "")
        ws.cell(i, 6).value = log.get("zone", "")

    # Write opponent logs
    if opp_logs:
        opp_start = len(logs) + 2
        ws.cell(opp_start, 1).value = "--- ADVERSARIO ---"
        ws.cell(opp_start, 1).font = openpyxlFont(bold=True)
        for i, opp in enumerate(opp_logs, start=opp_start + 1):
            ws.cell(i, 2).value = opp.get("min", 0)
            ws.cell(i, 5).value = opp.get("action", "")
            ws.cell(i, 6).value = opp.get("zone", "")

    wb.save(EXCEL_PATH)
    print(f"💾 Excel atualizado: {len(logs)} ações")

from openpyxl.styles import Font as openpyxlFont

def import_csv(csv_path):
    """Importa um ficheiro CSV para o Excel"""
    if not os.path.exists(csv_path):
        print(f"❌ CSV não encontrado: {csv_path}")
        return

    logs = []
    opp_logs = []
    is_opp = False

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        for row in reader:
            if len(row) == 0 or row[0] == '--- ACCOES ADVERSARIO ---':
                is_opp = True
                continue
            if len(row) < 6:
                continue
            if row[0] == 'Data':
                continue  # header row

            entry = {
                "min": row[1] if len(row) > 1 else 0,
                "player": row[2] if len(row) > 2 else "",
                "num": row[3] if len(row) > 3 else "",
                "action": row[4] if len(row) > 4 else "",
                "zone": row[5] if len(row) > 5 else ""
            }

            if is_opp:
                opp_logs.append({
                    "min": row[0],
                    "action": row[1],
                    "zone": row[2]
                })
            else:
                logs.append(entry)

    write_to_excel(logs, opp_logs)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--import", dest="csv_file", help="Importar CSV")
    parser.add_argument("--port", dest="port", type=int, default=5557)
    args = parser.parse_args()

    if args.csv_file:
        import_csv(args.csv_file)
    else:
        print(f"🚀 Server ativo em http://localhost:{args.port}")
        print(f"📊 Excel: {EXCEL_PATH}")
        server = HTTPServer(("localhost", args.port), SyncHandler)
        server.serve_forever()